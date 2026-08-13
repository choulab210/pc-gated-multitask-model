"""
Generate Supplementary Table S5
===============================

Table S5. External Validation Performance for Individual Nanoparticles.

Input
-----
results/external_validation/per_np_performance.csv

Expected columns
----------------
NP_ID
F1
Cosine
Observed_present
Predicted_present
TP
FP
FN

Outputs
-------
tables/Table_S5_external_validation_by_NP.csv
tables/Table_S5_external_validation_by_NP.xlsx

Run
---
python scripts/tables/tableS5_external_validation_by_np.py
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd


# ======================================================================
# Paths
# ======================================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]

INPUT_FILE = (
    PROJECT_ROOT
    / "results"
    / "external_validation"
    / "per_np_performance.csv"
)

TABLE_DIR = (
    PROJECT_ROOT
    / "tables"
)

TABLE_DIR.mkdir(
    parents=True,
    exist_ok=True,
)

CSV_FILE = (
    TABLE_DIR
    / "Table_S5_external_validation_by_NP.csv"
)

XLSX_FILE = (
    TABLE_DIR
    / "Table_S5_external_validation_by_NP.xlsx"
)


# ======================================================================
# Table text
# ======================================================================

TABLE_TITLE = (
    "Table S5. External Validation Performance for Individual Nanoparticles"
)

NOTE = (
    "Note: External validation was performed using eight independent "
    "NP–PC samples without model retraining. Evaluation was restricted "
    "to 124 proteins that were included in the trained model panel and "
    "detected in the external validation dataset. F1 summarizes adsorption "
    "classification performance for each nanoparticle, whereas cosine "
    "similarity quantifies agreement between predicted and observed "
    "protein corona abundance profiles. Observed proteins and predicted "
    "proteins indicate the numbers of proteins classified as present in "
    "the observed and predicted profiles, respectively. TP, FP, and FN "
    "denote true positives, false positives, and false negatives."
)


# ======================================================================
# Load and prepare data
# ======================================================================


def load_table_data() -> pd.DataFrame:
    """
    Load per-NP external-validation performance.
    """

    if not INPUT_FILE.exists():

        raise FileNotFoundError(
            "Per-NP external-validation file not found:\n"
            f"{INPUT_FILE}\n\n"
            "Run first:\n"
            "python scripts/run_external_validation.py"
        )

    dataframe = pd.read_csv(
        INPUT_FILE
    )

    required_columns = {
        "NP_ID",
        "F1",
        "Cosine",
        "Observed_present",
        "Predicted_present",
        "TP",
        "FP",
        "FN",
    }

    missing_columns = (
        required_columns
        - set(
            dataframe.columns
        )
    )

    if missing_columns:

        raise ValueError(
            "per_np_performance.csv is missing required column(s): "
            f"{sorted(missing_columns)}"
        )

    # ------------------------------------------------------------------
    # Keep only manuscript columns
    # ------------------------------------------------------------------

    dataframe = (
        dataframe[
            [
                "NP_ID",
                "F1",
                "Cosine",
                "Observed_present",
                "Predicted_present",
                "TP",
                "FP",
                "FN",
            ]
        ]
        .copy()
    )

    # ------------------------------------------------------------------
    # Rename columns for publication
    # ------------------------------------------------------------------

    dataframe = dataframe.rename(
        columns={
            "NP_ID":
                "External NP",

            "Cosine":
                "Cosine similarity",

            "Observed_present":
                "Observed proteins",

            "Predicted_present":
                "Predicted proteins",
        }
    )

    # ------------------------------------------------------------------
    # Sort NP01 -> NP08
    # ------------------------------------------------------------------

    dataframe[
        "External NP"
    ] = dataframe[
        "External NP"
    ].astype(str)

    dataframe = (
        dataframe
        .sort_values(
            "External NP"
        )
        .reset_index(
            drop=True
        )
    )

    # ------------------------------------------------------------------
    # Numeric formatting
    # ------------------------------------------------------------------

    dataframe[
        "F1"
    ] = (
        pd.to_numeric(
            dataframe[
                "F1"
            ],
            errors="raise",
        )
        .round(
            3
        )
    )

    dataframe[
        "Cosine similarity"
    ] = (
        pd.to_numeric(
            dataframe[
                "Cosine similarity"
            ],
            errors="raise",
        )
        .round(
            3
        )
    )

    integer_columns = [
        "Observed proteins",
        "Predicted proteins",
        "TP",
        "FP",
        "FN",
    ]

    for column in integer_columns:

        dataframe[
            column
        ] = (
            pd.to_numeric(
                dataframe[
                    column
                ],
                errors="raise",
            )
            .astype(
                int
            )
        )

    return dataframe


# ======================================================================
# Export CSV
# ======================================================================


def save_csv(
    dataframe: pd.DataFrame,
) -> None:

    dataframe.to_csv(
        CSV_FILE,
        index=False,
        float_format="%.3f",
    )


# ======================================================================
# Export Excel
# ======================================================================


def save_excel(
    dataframe: pd.DataFrame,
) -> None:

    from openpyxl import Workbook

    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = (
        "Table S5"
    )

    # ------------------------------------------------------------------
    # Fonts
    # ------------------------------------------------------------------

    title_font = Font(
        name="Arial",
        size=11,
        bold=True,
    )

    header_font = Font(
        name="Arial",
        size=10,
        bold=True,
    )

    body_font = Font(
        name="Arial",
        size=10,
    )

    note_font = Font(
        name="Arial",
        size=9,
    )

    # ------------------------------------------------------------------
    # Header fill/border
    # ------------------------------------------------------------------

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    thin_gray = Side(
        style="thin",
        color="B7B7B7",
    )

    header_border = Border(
        bottom=thin_gray,
    )

    # ==================================================================
    # Title
    # ==================================================================

    n_columns = len(
        dataframe.columns
    )

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=n_columns,
    )

    title_cell = worksheet.cell(
        row=1,
        column=1,
        value=TABLE_TITLE,
    )

    title_cell.font = (
        title_font
    )

    title_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    worksheet.row_dimensions[
        1
    ].height = 30

    # ==================================================================
    # Header
    # ==================================================================

    header_row = 3

    for column_index, header in enumerate(
        dataframe.columns,
        start=1,
    ):

        cell = worksheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )

        cell.font = (
            header_font
        )

        cell.fill = (
            header_fill
        )

        cell.border = (
            header_border
        )

        cell.alignment = Alignment(
            horizontal=(
                "left"
                if column_index == 1
                else "center"
            ),
            vertical="center",
            wrap_text=True,
        )

    # ==================================================================
    # Body
    # ==================================================================

    start_row = (
        header_row
        + 1
    )

    for row_offset, row in enumerate(
        dataframe.itertuples(
            index=False,
            name=None,
        )
    ):

        excel_row = (
            start_row
            + row_offset
        )

        for column_index, value in enumerate(
            row,
            start=1,
        ):

            cell = worksheet.cell(
                row=excel_row,
                column=column_index,
                value=value,
            )

            cell.font = (
                body_font
            )

            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if column_index == 1
                    else "center"
                ),
                vertical="center",
            )

        # F1
        worksheet.cell(
            row=excel_row,
            column=2,
        ).number_format = "0.000"

        # Cosine similarity
        worksheet.cell(
            row=excel_row,
            column=3,
        ).number_format = "0.000"

    # ==================================================================
    # Note
    # ==================================================================

    note_row = (
        start_row
        + len(
            dataframe
        )
        + 2
    )

    worksheet.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row,
        end_column=n_columns,
    )

    note_cell = worksheet.cell(
        row=note_row,
        column=1,
        value=NOTE,
    )

    note_cell.font = (
        note_font
    )

    note_cell.alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )

    worksheet.row_dimensions[
        note_row
    ].height = 85

    # ==================================================================
    # Column widths
    # ==================================================================

    widths = {
        "A": 14,
        "B": 10,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 8,
        "G": 8,
        "H": 8,
    }

    for column_letter, width in widths.items():

        worksheet.column_dimensions[
            column_letter
        ].width = width

    worksheet.freeze_panes = (
        "A4"
    )

    workbook.save(
        XLSX_FILE
    )


# ======================================================================
# Main
# ======================================================================


def main() -> None:

    print(
        "=" * 72
    )

    print(
        "GENERATING TABLE S5"
    )

    print(
        "=" * 72
    )

    dataframe = (
        load_table_data()
    )

    print()

    print(
        TABLE_TITLE
    )

    print()

    print(
        dataframe.to_string(
            index=False,
            formatters={
                "F1":
                    lambda value:
                        f"{value:.3f}",

                "Cosine similarity":
                    lambda value:
                        f"{value:.3f}",
            },
        )
    )

    save_csv(
        dataframe
    )

    save_excel(
        dataframe
    )

    print()

    print(
        "=" * 72
    )

    print(
        "TABLE S5 COMPLETE"
    )

    print(
        "=" * 72
    )

    print()

    print(
        "Saved:"
    )

    print(
        CSV_FILE
    )

    print(
        XLSX_FILE
    )


if __name__ == "__main__":
    main()