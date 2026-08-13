"""
Generate Supplementary Table S5
===============================

Table S5. External Validation Performance of the Gated Two-Head Model
by Protein Functional Category.

Input
-----
results/external_validation/category_performance.csv

Expected columns
----------------
Category
N
F1
Cosine

Outputs
-------
tables/Table_S5_external_validation.csv
tables/Table_S5_external_validation.xlsx

Run
---
python scripts/tables/tableS5_external_validation.py
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd


# ======================================================================
# Paths
# ======================================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]

INPUT_FILE = (
    PROJECT_ROOT
    / "results"
    / "external_validation"
    / "category_performance.csv"
)

TABLE_DIR = (
    PROJECT_ROOT
    / "tables"
)

TABLE_DIR.mkdir(
    parents=True,
    exist_ok=True,
)

CSV_FILE = (
    TABLE_DIR
    / "Table_S5_external_validation.csv"
)

XLSX_FILE = (
    TABLE_DIR
    / "Table_S5_external_validation.xlsx"
)


# ======================================================================
# Table configuration
# ======================================================================

TABLE_TITLE = (
    "Table S5. External Validation Performance of the "
    "Gated Two-Head Model by Protein Functional Category"
)

NOTE = (
    "Note: External validation was performed using eight independent "
    "NP–PC samples without model retraining. Evaluation was restricted "
    "to 124 proteins that were included in the model panel and detected "
    "in the external validation dataset. F1 score summarizes protein "
    "adsorption classification performance, whereas cosine similarity "
    "evaluates agreement between predicted and observed protein corona "
    "abundance profiles. Category-level metrics were calculated using "
    "proteins assigned to each functional category. N indicates the "
    "number of overlapping proteins evaluated within each category."
)


CATEGORY_ORDER = [
    "Overall",
    "Apolipoproteins",
    "Coagulation/Fibrinogen",
    "Complement System",
    "Cytoskeletal",
    "Immunoglobulins",
    "Metabolic Enzymes",
    "Protease/Inhibitors",
    "Transport/Binding",
    "Other/Mixed",
]


# ======================================================================
# Load and validate
# ======================================================================


def load_table_data() -> pd.DataFrame:
    """
    Load external-validation category performance and prepare Table S5.
    """

    if not INPUT_FILE.exists():

        raise FileNotFoundError(
            "External-validation category performance file not found:\n"
            f"{INPUT_FILE}\n\n"
            "Run first:\n"
            "python scripts/run_external_validation.py"
        )

    dataframe = pd.read_csv(
        INPUT_FILE
    )

    required_columns = {
        "Category",
        "N",
        "F1",
        "Cosine",
    }

    missing_columns = (
        required_columns
        - set(
            dataframe.columns
        )
    )

    if missing_columns:

        raise ValueError(
            "category_performance.csv is missing required column(s): "
            f"{sorted(missing_columns)}"
        )

    # ------------------------------------------------------------------
    # Retain only required columns
    # ------------------------------------------------------------------

    dataframe = (
        dataframe[
            [
                "Category",
                "N",
                "F1",
                "Cosine",
            ]
        ]
        .copy()
    )

    # ------------------------------------------------------------------
    # Enforce manuscript category order
    # ------------------------------------------------------------------

    dataframe[
        "Category"
    ] = pd.Categorical(
        dataframe[
            "Category"
        ],
        categories=CATEGORY_ORDER,
        ordered=True,
    )

    dataframe = (
        dataframe
        .sort_values(
            "Category"
        )
        .reset_index(
            drop=True
        )
    )

    if dataframe[
        "Category"
    ].isna().any():

        raise ValueError(
            "An unrecognized protein category was found in "
            "category_performance.csv."
        )

    # ------------------------------------------------------------------
    # Rename for manuscript presentation
    # ------------------------------------------------------------------

    dataframe = dataframe.rename(
        columns={
            "Category":
                "Protein functional category",

            "N":
                "N",

            "F1":
                "F1",

            "Cosine":
                "Cosine similarity",
        }
    )

    # ------------------------------------------------------------------
    # Three decimal places for performance metrics
    # ------------------------------------------------------------------

    dataframe[
        "F1"
    ] = (
        pd.to_numeric(
            dataframe[
                "F1"
            ],
            errors="raise",
        )
        .round(
            3
        )
    )

    dataframe[
        "Cosine similarity"
    ] = (
        pd.to_numeric(
            dataframe[
                "Cosine similarity"
            ],
            errors="raise",
        )
        .round(
            3
        )
    )

    dataframe[
        "N"
    ] = (
        pd.to_numeric(
            dataframe[
                "N"
            ],
            errors="raise",
        )
        .astype(
            int
        )
    )

    return dataframe


# ======================================================================
# Export
# ======================================================================


def save_csv(
    dataframe: pd.DataFrame,
) -> None:
    """
    Save machine-readable CSV table.
    """

    dataframe.to_csv(
        CSV_FILE,
        index=False,
        float_format="%.3f",
    )


def save_excel(
    dataframe: pd.DataFrame,
) -> None:
    """
    Save publication-ready Excel version of Table S5.
    """

    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Font,
        PatternFill,
        Border,
        Side,
    )

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = (
        "Table S5"
    )

    # ------------------------------------------------------------------
    # Styles
    # ------------------------------------------------------------------

    font_name = "Arial"

    title_font = Font(
        name=font_name,
        size=11,
        bold=True,
    )

    header_font = Font(
        name=font_name,
        size=10,
        bold=True,
    )

    body_font = Font(
        name=font_name,
        size=10,
    )

    overall_font = Font(
        name=font_name,
        size=10,
        bold=True,
    )

    note_font = Font(
        name=font_name,
        size=9,
        italic=False,
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    thin_gray = Side(
        style="thin",
        color="B7B7B7",
    )

    bottom_border = Border(
        bottom=thin_gray,
    )

    # ------------------------------------------------------------------
    # Title
    # ------------------------------------------------------------------

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=4,
    )

    title_cell = worksheet.cell(
        row=1,
        column=1,
        value=TABLE_TITLE,
    )

    title_cell.font = (
        title_font
    )

    title_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    worksheet.row_dimensions[
        1
    ].height = 32

    # ------------------------------------------------------------------
    # Header
    # ------------------------------------------------------------------

    header_row = 3

    headers = [
        "Protein functional category",
        "N",
        "F1",
        "Cosine similarity",
    ]

    for column_index, header in enumerate(
        headers,
        start=1,
    ):

        cell = worksheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )

        cell.font = (
            header_font
        )

        cell.fill = (
            header_fill
        )

        cell.alignment = Alignment(
            horizontal=(
                "left"
                if column_index == 1
                else "center"
            ),
            vertical="center",
            wrap_text=True,
        )

        cell.border = (
            bottom_border
        )

    # ------------------------------------------------------------------
    # Body
    # ------------------------------------------------------------------

    start_row = (
        header_row
        + 1
    )

    for row_offset, row in enumerate(
        dataframe.itertuples(
            index=False,
            name=None,
        )
    ):

        excel_row = (
            start_row
            + row_offset
        )

        is_overall = (
            str(
                row[
                    0
                ]
            )
            == "Overall"
        )

        for column_index, value in enumerate(
            row,
            start=1,
        ):

            cell = worksheet.cell(
                row=excel_row,
                column=column_index,
                value=value,
            )

            cell.font = (
                overall_font
                if is_overall
                else body_font
            )

            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if column_index == 1
                    else "center"
                ),
                vertical="center",
            )

        # Format metrics to 3 decimal places.
        worksheet.cell(
            row=excel_row,
            column=3,
        ).number_format = "0.000"

        worksheet.cell(
            row=excel_row,
            column=4,
        ).number_format = "0.000"

    # ------------------------------------------------------------------
    # Note
    # ------------------------------------------------------------------

    note_row = (
        start_row
        + len(
            dataframe
        )
        + 2
    )

    worksheet.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row,
        end_column=4,
    )

    note_cell = worksheet.cell(
        row=note_row,
        column=1,
        value=NOTE,
    )

    note_cell.font = (
        note_font
    )

    note_cell.alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )

    worksheet.row_dimensions[
        note_row
    ].height = 75

    # ------------------------------------------------------------------
    # Column widths
    # ------------------------------------------------------------------

    worksheet.column_dimensions[
        "A"
    ].width = 32

    worksheet.column_dimensions[
        "B"
    ].width = 10

    worksheet.column_dimensions[
        "C"
    ].width = 12

    worksheet.column_dimensions[
        "D"
    ].width = 20

    # ------------------------------------------------------------------
    # Freeze header
    # ------------------------------------------------------------------

    worksheet.freeze_panes = (
        "A4"
    )

    workbook.save(
        XLSX_FILE
    )


# ======================================================================
# Main
# ======================================================================


def main() -> None:

    print(
        "=" * 72
    )

    print(
        "GENERATING TABLE S5"
    )

    print(
        "=" * 72
    )

    dataframe = (
        load_table_data()
    )

    print()

    print(
        TABLE_TITLE
    )

    print()

    print(
        dataframe.to_string(
            index=False,
            formatters={
                "F1":
                    lambda value:
                        f"{value:.3f}",

                "Cosine similarity":
                    lambda value:
                        f"{value:.3f}",
            },
        )
    )

    save_csv(
        dataframe
    )

    save_excel(
        dataframe
    )

    print()

    print(
        "=" * 72
    )

    print(
        "TABLE S5 COMPLETE"
    )

    print(
        "=" * 72
    )

    print()

    print(
        "Saved:"
    )

    print(
        CSV_FILE
    )

    print(
        XLSX_FILE
    )


if __name__ == "__main__":
    main()